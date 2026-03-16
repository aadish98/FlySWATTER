"""Callable score-analysis service used by both CLI and GUI flows."""

from __future__ import annotations

import math
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Dict, List, Optional

import matplotlib

matplotlib.use("Agg")

import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from matplotlib.artist import setp as _mpl_setp
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from matplotlib.patches import Patch
from matplotlib.ticker import FuncFormatter, MultipleLocator, NullFormatter

from ScoreArousability import (
    _ensure_numeric,
    _parse_start_datetime_from_filename,
    _read_zantiks_file,
    detect_pulses_by_row,
    load_mapping,
    sanitize_excel_sheet_name,
    score_genotype_timerows,
)
from services.models import ScoreAnalysisResult
from services.output_packaging import create_zip_from_paths
from services.power_management import prevent_sleep

ProgressCallback = Optional[Callable[[int, str], None]]


def run_score_analysis_from_mapping_file(
    input_file: str | Path,
    mapping_file: str | Path,
    output_dir: str | Path,
    *,
    pre_sec: int = 300,
    post_sec: int = 120,
    max_pulses: int = 0,
    dry_run: bool = False,
    progress_callback: ProgressCallback = None,
) -> ScoreAnalysisResult:
    mapping, genotype_order = load_mapping(str(mapping_file))
    return run_score_analysis(
        input_file,
        mapping,
        genotype_order,
        output_dir,
        pre_sec=pre_sec,
        post_sec=post_sec,
        max_pulses=max_pulses,
        dry_run=dry_run,
        progress_callback=progress_callback,
    )


def run_score_analysis(
    input_file: str | Path,
    mapping: Dict[str, List[str]],
    genotype_order: List[str],
    output_dir: str | Path,
    *,
    pre_sec: int = 300,
    post_sec: int = 120,
    sleep_threshold_sec: Optional[int] = None,
    max_pulses: int = 0,
    dry_run: bool = False,
    progress_callback: ProgressCallback = None,
) -> ScoreAnalysisResult:
    input_path = Path(input_file)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    sleep_dir = output_path / "Sleep"
    sleep_dir.mkdir(parents=True, exist_ok=True)

    mapping = {genotype: list(dict.fromkeys(wells)) for genotype, wells in mapping.items()}
    genotype_order = list(genotype_order)
    resolved_sleep_threshold_sec = int(sleep_threshold_sec) if sleep_threshold_sec is not None else int(pre_sec)
    if resolved_sleep_threshold_sec <= 0:
        resolved_sleep_threshold_sec = 300

    with prevent_sleep():
        _emit(progress_callback, 5, "Loading Zantiks behavior data...")
        df = _read_zantiks_file(str(input_path))
        filename_start_dt = _parse_start_datetime_from_filename(str(input_path))
        base = input_path.stem

        required = ["DAY", "HOUR", "TIME_BIN", "VIB", "RUNTIME"]
        for column in required:
            if column not in df.columns:
                raise KeyError(f"Required column '{column}' not in data")
        _ensure_numeric(df, ["TIME_BIN", "DAY", "HOUR", "VIB", "RUNTIME"])
        df["VIB"] = df["VIB"].fillna(0)

        base_day = int(pd.to_numeric(df["DAY"], errors="coerce").min())
        day_val = pd.to_numeric(df["DAY"], errors="coerce")
        hour_val = pd.to_numeric(df["HOUR"], errors="coerce")
        timebin_val = pd.to_numeric(df["TIME_BIN"], errors="coerce")
        abs_seconds = (day_val - base_day) * 24 * 3600 + hour_val * 3600 + timebin_val
        abs_seconds = abs_seconds.replace([float("inf"), float("-inf")], pd.NA).fillna(0)
        df["ABS_BIN"] = abs_seconds.astype(float).round().astype(int)

        pulses = detect_pulses_by_row(df)
        if max_pulses and max_pulses > 0:
            pulses = pulses[:max_pulses]

        _emit(progress_callback, 18, f"Scoring {len(pulses)} pulse(s) across {len(genotype_order)} genotype(s)...")
        compute_pulse_zt = _make_pulse_zt_computer(df, filename_start_dt)
        out_df = _build_score_table(
            df,
            pulses,
            mapping,
            genotype_order,
            pre_sec,
            post_sec,
            dry_run,
            compute_pulse_zt,
        )

        _emit(progress_callback, 38, "Writing score tables...")
        hour_runtime = _build_hour_runtime(df)
        pulse_regimen_df = _build_pulse_regimen(pulses, compute_pulse_zt, filename_start_dt)
        arousal_workbook = output_path / f"ArousalScores_{base}.xlsx"
        with pd.ExcelWriter(arousal_workbook, engine="openpyxl") as writer:
            out_df.to_excel(writer, index=False, sheet_name="Scores")
            hour_runtime.to_excel(writer, index=False, sheet_name="HourRuntime")
            pulse_regimen_df.to_excel(writer, index=False, sheet_name="Pulse Regimen")

        _emit(progress_callback, 52, "Rendering arousal plots...")
        arousal_plot_paths = _plot_aggregated_arousal(out_df, genotype_order, output_path, base)
        protocol_plot = _plot_protocol(df, pulses, filename_start_dt, compute_pulse_zt, output_path, base)

        _emit(progress_callback, 72, "Generating sleep outputs...")
        sleep_outputs = _generate_sleep_outputs(
            df,
            mapping,
            genotype_order,
            filename_start_dt,
            compute_pulse_zt,
            sleep_dir,
            base,
            sleep_threshold_sec=resolved_sleep_threshold_sec,
        )

        _emit(progress_callback, 92, "Packaging download bundles...")
        arousal_zip = output_path / f"Arousal_Data_{base}.zip"
        sleep_zip = output_path / f"Sleep_Data_{base}.zip"
        arousal_paths = [arousal_workbook, protocol_plot, *arousal_plot_paths]
        sleep_paths = [
            sleep_outputs["sleep_totals_workbook"],
            sleep_outputs["sleep_profile_plot"],
            sleep_outputs["sleep_pct_plot"],
            sleep_outputs["individual_dir"],
        ]
        create_zip_from_paths(arousal_zip, [path for path in arousal_paths if path], base_dir=output_path)
        create_zip_from_paths(sleep_zip, [path for path in sleep_paths if path], base_dir=sleep_dir)

    _emit(progress_callback, 100, "Score analysis complete.")
    genotype_counts = {genotype: len(mapping.get(genotype, [])) for genotype in genotype_order}
    preview_paths = [path for path in [protocol_plot, arousal_plot_paths[0] if arousal_plot_paths else None, sleep_outputs["sleep_profile_plot"]] if path]
    return ScoreAnalysisResult(
        output_dir=output_path,
        arousal_workbook=arousal_workbook,
        arousal_plot_paths=arousal_plot_paths,
        protocol_plot=protocol_plot,
        sleep_totals_workbook=sleep_outputs["sleep_totals_workbook"],
        sleep_profile_plot=sleep_outputs["sleep_profile_plot"],
        sleep_pct_plot=sleep_outputs["sleep_pct_plot"],
        sleep_individual_dir=sleep_outputs["individual_dir"],
        arousal_zip=arousal_zip,
        sleep_zip=sleep_zip,
        preview_plot_paths=preview_paths,
        genotype_counts=genotype_counts,
    )


def _emit(callback: ProgressCallback, value: int, message: str) -> None:
    if callback is not None:
        callback(value, message)


def _make_pulse_zt_computer(df: pd.DataFrame, filename_start_dt: Optional[datetime]) -> Callable[[float, Optional[int], Optional[int]], Optional[int]]:
    def _compute_pulse_zt(start_runtime, p_hour, pulse_start_row):
        if filename_start_dt is not None and pd.notna(start_runtime):
            pulse_wc = filename_start_dt + timedelta(seconds=float(start_runtime))
            rounded_hour = (pulse_wc.hour + 1) % 24 if pulse_wc.minute >= 30 else pulse_wc.hour
            return (rounded_hour - 8) % 24
        if pd.notna(p_hour):
            hour = int(p_hour)
            if pulse_start_row is not None and "TIME_BIN" in df.columns:
                sr = int(pulse_start_row)
                if 0 <= sr < len(df):
                    tb = pd.to_numeric(df.iloc[sr].get("TIME_BIN", None), errors="coerce")
                    if pd.notna(tb) and float(tb) >= 1800:
                        hour = (hour + 1) % 24
            return (hour - 8) % 24
        return None

    return _compute_pulse_zt


def _build_score_table(
    df: pd.DataFrame,
    pulses,
    mapping: Dict[str, List[str]],
    genotype_order: List[str],
    pre_sec: int,
    post_sec: int,
    dry_run: bool,
    compute_pulse_zt: Callable[[float, Optional[int], Optional[int]], Optional[int]],
) -> pd.DataFrame:
    rows = []
    for pulse_i, (start_row, end_row, p_day, p_hour, p_int_temp1, start_runtime) in enumerate(pulses, start=1):
        duration = int(end_row - start_row + 1)
        day_value = int(p_day) if p_day is not None else None
        hour_value = int(p_hour) if p_hour is not None else None
        int_temp1_value = float(p_int_temp1) if pd.notna(p_int_temp1) else None
        zt_value = compute_pulse_zt(start_runtime, p_hour, start_row)
        for genotype in genotype_order:
            wells = mapping.get(genotype, [])
            result = score_genotype_timerows(df, wells, start_row, end_row, pre_sec, post_sec)
            if result is None:
                if not dry_run:
                    continue
                pre_window = (max(start_row - pre_sec, 0), max(start_row - 1, 0))
                arousal_window = (min(end_row + 1, len(df) - 1), min(end_row + post_sec, len(df) - 1))
                asleep_wells = list(wells)
                awoken_wells: List[str] = []
                n_asleep = len(asleep_wells)
                n_awoken = 0
                pct = 0.0
            else:
                n_asleep, asleep_wells, n_awoken, awoken_wells, pct, pre_window, arousal_window = result

            rows.append(
                {
                    "Pulse #": pulse_i,
                    "Pulse Start Row": start_row,
                    "Pulse End Row": end_row,
                    "Pulse Start Runtime": start_runtime,
                    "Pulse Duration (s)": duration,
                    "Pulse Day": day_value,
                    "Pulse Hour": hour_value,
                    "ZT": zt_value,
                    "Pulse INT_TEMP1": int_temp1_value,
                    "Genotype": genotype,
                    "Asleep (n)": n_asleep,
                    "Asleep Wells": ",".join(asleep_wells),
                    "Awoken (n)": n_awoken,
                    "Awoken Wells": ",".join(awoken_wells),
                    "% Awoken": pct,
                    "Pre Window (rows)": f"{pre_window[0]}-{pre_window[1]}",
                    "Arousal Window (rows)": f"{arousal_window[0]}-{arousal_window[1]}",
                }
            )
    return pd.DataFrame(
        rows,
        columns=[
            "Pulse #",
            "Pulse Start Row",
            "Pulse End Row",
            "Pulse Start Runtime",
            "Pulse Duration (s)",
            "Pulse Day",
            "Pulse Hour",
            "ZT",
            "Pulse INT_TEMP1",
            "Genotype",
            "Asleep (n)",
            "Asleep Wells",
            "Awoken (n)",
            "Awoken Wells",
            "% Awoken",
            "Pre Window (rows)",
            "Arousal Window (rows)",
        ],
    )


def _build_hour_runtime(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["DAY", "HOUR"] + (["L_OR_D"] if "L_OR_D" in df.columns else [])
    hour_df = df[group_cols + ["RUNTIME"]].copy()
    hour_df = hour_df.dropna(subset=group_cols + ["RUNTIME"]).copy()
    hour_df["DAY"] = pd.to_numeric(hour_df["DAY"], errors="coerce")
    hour_df["HOUR"] = pd.to_numeric(hour_df["HOUR"], errors="coerce")
    hour_df["RUNTIME"] = pd.to_numeric(hour_df["RUNTIME"], errors="coerce")
    hour_df = hour_df.dropna(subset=["DAY", "HOUR", "RUNTIME"]).copy()
    hour_df["DAY"] = hour_df["DAY"].astype(int)
    hour_df["HOUR"] = hour_df["HOUR"].astype(int)
    hour_agg = (
        hour_df.groupby(group_cols, sort=False)["RUNTIME"]
        .agg(min_runtime="min", max_runtime="max")
        .reset_index()
    )
    duration_s = (hour_agg["max_runtime"] - hour_agg["min_runtime"]).clip(lower=0).round().astype(int)
    hour_runtime = hour_agg[group_cols + ["min_runtime", "max_runtime"]].copy()
    insert_pos = len(group_cols)
    hour_runtime.insert(insert_pos, "Min RUNTIME", hour_runtime.pop("min_runtime"))
    hour_runtime.insert(insert_pos + 1, "Max RUNTIME", hour_runtime.pop("max_runtime"))
    hour_runtime["Duration (s)"] = duration_s
    hour_runtime = hour_runtime.sort_values(group_cols).reset_index(drop=True)
    hour_runtime["Duration (min)"] = hour_runtime["Duration (s)"] / 60.0
    hour_runtime["Duration (hr)"] = hour_runtime["Duration (s)"] / 3600.0
    return hour_runtime


def _build_pulse_regimen(pulses, compute_pulse_zt, filename_start_dt: Optional[datetime]) -> pd.DataFrame:
    rows = []
    for pulse_i, (start_row, end_row, p_day, p_hour, p_int_temp1, start_runtime) in enumerate(pulses, start=1):
        duration = int(end_row - start_row + 1)
        if filename_start_dt is not None and pd.notna(start_runtime):
            pulse_datetime = filename_start_dt + timedelta(seconds=float(start_runtime))
            time_str = pulse_datetime.strftime("%I:%M %p").lstrip("0")
            date_str = pulse_datetime.strftime("%Y-%m-%d")
        else:
            time_str = None
            date_str = f"Day {int(p_day)}" if pd.notna(p_day) else None
        rows.append(
            {
                "Pulse #": pulse_i,
                "Experiment Day": int(p_day) if pd.notna(p_day) else None,
                "Date": date_str,
                "Time Delivered": time_str,
                "ZT": compute_pulse_zt(start_runtime, p_hour, start_row),
                "Runtime (s)": start_runtime if pd.notna(start_runtime) else None,
                "Pulse Duration (s)": duration,
                "Temperature (°C)": float(p_int_temp1) if pd.notna(p_int_temp1) else None,
            }
        )
    return pd.DataFrame(
        rows,
        columns=[
            "Pulse #",
            "Experiment Day",
            "Date",
            "Time Delivered",
            "ZT",
            "Runtime (s)",
            "Pulse Duration (s)",
            "Temperature (°C)",
        ],
    )


def _plot_aggregated_arousal(out_df: pd.DataFrame, genotype_order: List[str], output_dir: Path, base: str) -> List[Path]:
    if out_df.empty:
        return []
    plot_df = out_df.dropna(subset=["Pulse Day", "Pulse Hour"]).copy()
    if plot_df.empty:
        return []
    pulse_labels = (
        plot_df.groupby(["Pulse Day", "Pulse #"], sort=False)
        .first()
        .reset_index()[["Pulse Day", "Pulse #", "ZT"]]
    )
    pulse_labels["Pulse Label"] = pulse_labels.apply(lambda row: f"D{int(row['Pulse Day'])}-P{int(row['Pulse #'])}", axis=1)
    pulse_labels["Label"] = pulse_labels.apply(
        lambda row: f"D{int(row['Pulse Day'])}-P{int(row['Pulse #'])}\nZT{int(round(row['ZT']))}" if pd.notna(row["ZT"]) else f"D{int(row['Pulse Day'])}-P{int(row['Pulse #'])}",
        axis=1,
    )
    plot_df["Pulse Label"] = plot_df.apply(lambda row: f"D{int(row['Pulse Day'])}-P{int(row['Pulse #'])}", axis=1)
    pulse_data = plot_df.groupby(["Pulse Label", "Genotype"])["% Awoken"].mean().unstack(fill_value=0)
    asleep_counts = plot_df.groupby(["Pulse Label", "Genotype"])["Asleep (n)"].first().unstack(fill_value=0)
    awoken_counts = plot_df.groupby(["Pulse Label", "Genotype"])["Awoken (n)"].first().unstack(fill_value=0)

    fig = Figure(figsize=(14, 8))
    FigureCanvasAgg(fig)
    ax = fig.subplots()
    pulse_index_labels = list(pulse_data.index)
    x = np.arange(len(pulse_index_labels))
    genotypes = [g for g in genotype_order if g in pulse_data.columns]
    width = 0.8 / max(len(genotypes), 1)
    for idx, genotype in enumerate(genotypes):
        values = pulse_data[genotype].values
        positions = x + (idx - len(genotypes) / 2 + 0.5) * width
        bars = ax.bar(positions, values, width, label=genotype, alpha=0.85)
        for bar, label in zip(bars, pulse_index_labels):
            awoken = int(awoken_counts.loc[label, genotype]) if genotype in awoken_counts.columns else 0
            asleep = int(asleep_counts.loc[label, genotype]) if genotype in asleep_counts.columns else 0
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                bar.get_height() if bar.get_height() > 0 else 0,
                f"{awoken}/{asleep}",
                ha="center",
                va="bottom",
                fontsize=7,
            )
    label_lookup = dict(zip(pulse_labels["Pulse Label"], pulse_labels["Label"]))
    pretty_labels = [label_lookup.get(label, label) for label in pulse_index_labels]
    ax.set_xticks(x)
    ax.set_xticklabels(pretty_labels)
    ax.set_ylim(0, 100)
    ax.set_xlabel("Pulse")
    ax.set_ylabel("% Awoken (awoken/asleep x 100%)")
    ax.set_title("Aggregated Arousal by Genotype and Pulse")
    ax.grid(True, alpha=0.3, axis="y")
    ax.legend(title="Genotype", bbox_to_anchor=(1.02, 1), loc="upper left")
    fig.tight_layout()
    fig_path = output_dir / f"Arousal_BarPlot_{base}.png"
    fig.savefig(fig_path, dpi=300, bbox_inches="tight")
    return [fig_path]


def _plot_protocol(df: pd.DataFrame, pulses, filename_start_dt: Optional[datetime], compute_pulse_zt, output_dir: Path, base: str) -> Optional[Path]:
    if filename_start_dt is None:
        return None
    missing_cols = [column for column in ["RUNTIME", "INT_TEMP1", "L_OR_D"] if column not in df.columns]
    if missing_cols:
        return None

    ld_df = df[["RUNTIME", "INT_TEMP1", "L_OR_D"]].copy()
    ld_df["RUNTIME"] = pd.to_numeric(ld_df["RUNTIME"], errors="coerce")
    ld_df["INT_TEMP1"] = pd.to_numeric(ld_df["INT_TEMP1"], errors="coerce")
    ld_df = ld_df.dropna(subset=["RUNTIME"]).sort_values("RUNTIME")
    ld_df["is_light"] = ld_df["L_OR_D"].apply(_to_is_light)
    ld_df["wall_clock"] = filename_start_dt + pd.to_timedelta(ld_df["RUNTIME"], unit="s")

    fig = Figure(figsize=(14, 8))
    FigureCanvasAgg(fig)
    ax1 = fig.subplots()
    if ld_df["INT_TEMP1"].notna().any():
        ax1.plot(ld_df["wall_clock"], ld_df["INT_TEMP1"], color="tab:red", linewidth=1.8, label="Internal Temp (°C)")
        ax1.yaxis.set_major_locator(MultipleLocator(1))
    ax1.set_ylabel("Temperature (°C)")

    times = ld_df["wall_clock"].tolist()
    states = ld_df["is_light"].tolist()
    if times:
        if len(times) > 1:
            ax1.set_xlim(times[0], times[-1])
            first_day = datetime.combine(times[0].date(), datetime.min.time())
            last_day = datetime.combine(times[-1].date(), datetime.min.time())
            day_cursor = first_day
            while day_cursor <= last_day:
                if times[0] <= day_cursor <= times[-1]:
                    ax1.axvline(day_cursor, color="black", linestyle="--", linewidth=1.0, alpha=0.35, zorder=1)
                day_cursor += timedelta(days=1)
        seg_start = times[0]
        seg_state = states[0]
        for idx in range(1, len(times)):
            if states[idx] != seg_state:
                if seg_state is not None:
                    ax1.axvspan(seg_start, times[idx], facecolor="#ffe066" if seg_state else "#001f3f", alpha=0.18, zorder=0)
                seg_start = times[idx]
                seg_state = states[idx]
        if seg_state is not None:
            ax1.axvspan(seg_start, times[-1], facecolor="#ffe066" if seg_state else "#001f3f", alpha=0.18, zorder=0)

    x0_num, x1_num = ax1.get_xlim()
    for pulse_i, (start_row, _end_row, _day, p_hour, _temp, start_runtime) in enumerate(pulses, start=1):
        if pd.isna(start_runtime):
            continue
        pulse_x = filename_start_dt + pd.to_timedelta(float(start_runtime), unit="s")
        pulse_zt = compute_pulse_zt(start_runtime, p_hour, start_row)
        label_parts = [f"P#{pulse_i}", pulse_x.strftime("%I:%M %p").lstrip("0")]
        if pulse_zt is not None:
            label_parts.append(f"ZT{int(pulse_zt)}")
        px_num = mdates.date2num(pulse_x)
        frac = (px_num - x0_num) / (x1_num - x0_num) if (x1_num - x0_num) else 0.5
        dx = 12 if frac < 0.08 else (-12 if frac > 0.92 else 0)
        dy = [-50, -70, -90][(pulse_i - 1) % 3]
        ax1.annotate(
            "\n".join(label_parts),
            xy=(pulse_x, -0.02),
            xycoords=("data", "axes fraction"),
            xytext=(dx, dy),
            textcoords="offset points",
            ha="center",
            va="top",
            fontsize=8,
            clip_on=False,
            arrowprops=dict(arrowstyle="-|>", lw=0.8, color="black"),
        )

    _apply_wall_clock_xaxis(ax1)
    ax1.legend(
        handles=[
            Patch(facecolor="#ffe066", edgecolor="none", alpha=0.35, label="Day (lights on)"),
            Patch(facecolor="#001f3f", edgecolor="none", alpha=0.35, label="Night (lights off)"),
        ],
        loc="upper right",
        bbox_to_anchor=(1.0, -0.28),
        borderaxespad=0.0,
        fontsize=8,
        title="Light Cycle",
        title_fontsize=8,
        framealpha=0.9,
    )
    ax1.set_xlabel("Wall Clock Time")
    ax1.set_title("Temperature and Light/Dark Protocol Over Time")
    ax1.grid(True, axis="x", which="major", alpha=0.4)
    ax1.grid(True, axis="x", which="minor", alpha=0.2)
    fig.tight_layout(rect=[0, 0.32, 1, 0.96])
    fig_path = output_dir / f"Temp_Light_WallClock_{base}.png"
    fig.savefig(fig_path, dpi=300, bbox_inches="tight")
    return fig_path


def _generate_sleep_outputs(
    df: pd.DataFrame,
    mapping: Dict[str, List[str]],
    genotype_order: List[str],
    filename_start_dt: Optional[datetime],
    compute_pulse_zt: Callable[[float, Optional[int], Optional[int]], Optional[int]],
    sleep_dir: Path,
    base: str,
    *,
    sleep_threshold_sec: int = 300,
) -> Dict[str, Optional[Path]]:
    sleep_threshold_sec = max(int(sleep_threshold_sec), 1)
    sleep_bin_sec = 1800
    all_mapped_wells = []
    for genotype in genotype_order:
        for well in mapping.get(genotype, []):
            if well not in all_mapped_wells:
                all_mapped_wells.append(well)
    valid_sleep_wells = [well for well in all_mapped_wells if well in df.columns]
    if not valid_sleep_wells:
        return {
            "sleep_totals_workbook": None,
            "sleep_profile_plot": None,
            "sleep_pct_plot": None,
            "individual_dir": None,
        }

    runtime_sleep = pd.to_numeric(df["RUNTIME"], errors="coerce")
    rt_valid_mask = runtime_sleep.notna()
    if rt_valid_mask.sum() == 0:
        return {
            "sleep_totals_workbook": None,
            "sleep_profile_plot": None,
            "sleep_pct_plot": None,
            "individual_dir": None,
        }

    rt_min_val = float(runtime_sleep[rt_valid_mask].min())
    sleep_per_well = {}
    for well in valid_sleep_wells:
        act = pd.to_numeric(df[well], errors="coerce").fillna(0).values
        sleep_per_well[well] = _compute_sleep_binary(act, threshold=sleep_threshold_sec)

    sleep_totals_book = _write_sleep_totals_workbook(
        df,
        mapping,
        genotype_order,
        filename_start_dt,
        sleep_per_well,
        sleep_dir,
        base,
    )

    bin_index = ((runtime_sleep - rt_min_val) // sleep_bin_sec).copy()
    bin_index[~rt_valid_mask] = -1
    bin_index = bin_index.astype(int)
    bin_start_runtime = bin_index * sleep_bin_sec + rt_min_val
    sleep_frame = pd.DataFrame(sleep_per_well, index=df.index)
    sleep_frame["bin_rt"] = bin_start_runtime
    sleep_frame = sleep_frame[bin_index >= 0].copy()

    binned_sleep_minutes = sleep_frame.groupby("bin_rt")[valid_sleep_wells].sum() / 60.0
    binned_sleep = binned_sleep_minutes * (3600.0 / sleep_bin_sec)
    bin_rt_values = binned_sleep.index.values
    if len(bin_rt_values) == 0:
        return {
            "sleep_totals_workbook": sleep_totals_book,
            "sleep_profile_plot": None,
            "sleep_pct_plot": None,
            "individual_dir": None,
        }

    use_wc_sleep = filename_start_dt is not None
    anchor_sleep = filename_start_dt
    sleep_x_vals = (
        [anchor_sleep + timedelta(seconds=float(runtime)) for runtime in bin_rt_values]
        if use_wc_sleep and anchor_sleep is not None
        else list(bin_rt_values / 3600.0)
    )

    ld_transitions = _collect_light_dark_segments(df, anchor_sleep, use_wc_sleep)
    sleep_pulses = detect_pulses_by_row(df)

    individual_dir = sleep_dir / "Individual Fly Sleep Profiles"
    individual_dir.mkdir(parents=True, exist_ok=True)
    for genotype in genotype_order:
        safe_genotype = genotype.replace("/", "_").replace(" ", "_").replace(";", "_")
        genotype_dir = individual_dir / safe_genotype
        genotype_dir.mkdir(parents=True, exist_ok=True)
        wells = [well for well in mapping.get(genotype, []) if well in binned_sleep.columns]
        for well in wells:
            fig = Figure(figsize=(14, 5))
            FigureCanvasAgg(fig)
            ax = fig.subplots()
            _add_day_night_shading(ax, ld_transitions, use_wc_sleep)
            y_data = binned_sleep[well].values
            ax.plot(sleep_x_vals, y_data, color="black", linewidth=1.0)
            ax.fill_between(sleep_x_vals, 0, y_data, alpha=0.3, color="steelblue")
            ax.set_ylabel("Sleep (min/hour)")
            ax.set_ylim(0, 60)
            ax.set_title(f"Sleep Profile - {genotype} - Well {well}")
            ax.grid(True, axis="y", alpha=0.3)
            ax.grid(True, axis="x", which="major", alpha=0.4, linewidth=1.2)
            ax.grid(True, axis="x", which="minor", alpha=0.2, linewidth=0.6)
            _format_sleep_xaxis(ax, use_wc_sleep)
            _add_pulse_arrows(ax, sleep_pulses, anchor_sleep, use_wc_sleep, compute_pulse_zt)
            fig.tight_layout(rect=[0, 0.15, 1, 0.96])
            fig_path = genotype_dir / f"Sleep_{well}_{base}.png"
            fig.savefig(fig_path, dpi=200, bbox_inches="tight")

    combined_series = []
    for genotype in genotype_order:
        wells = [well for well in mapping.get(genotype, []) if well in binned_sleep.columns]
        if not wells:
            continue
        geno_sleep = binned_sleep[wells]
        combined_series.append((genotype, len(wells), geno_sleep.mean(axis=1).values))

    sleep_profile_plot = None
    sleep_pct_plot = None
    if combined_series:
        fig = Figure(figsize=(14, 6))
        FigureCanvasAgg(fig)
        ax = fig.subplots()
        _add_day_night_shading(ax, ld_transitions, use_wc_sleep)
        color_map = matplotlib.colormaps["tab10"]
        for idx, (genotype, n_flies, mean_sleep) in enumerate(combined_series):
            ax.plot(sleep_x_vals, mean_sleep, linewidth=2.0, color=color_map(idx % 10), label=f"{genotype} (n={n_flies})")
        ax.set_ylabel("Sleep (min/hour)")
        ax.set_ylim(0, 60)
        ax.set_title("Sleep Profile - All Genotypes (mean)")
        ax.grid(True, axis="y", alpha=0.3)
        ax.grid(True, axis="x", which="major", alpha=0.4, linewidth=1.2)
        ax.grid(True, axis="x", which="minor", alpha=0.2, linewidth=0.6)
        _format_sleep_xaxis(ax, use_wc_sleep)
        _add_pulse_arrows(ax, sleep_pulses, anchor_sleep, use_wc_sleep, compute_pulse_zt)
        ax.legend(
            title="Genotype",
            loc="center left",
            bbox_to_anchor=(1.01, 0.5),
            fontsize=8,
            title_fontsize=8,
            framealpha=0.9,
            borderpad=0.25,
            labelspacing=0.25,
            handlelength=1.4,
            columnspacing=0.8,
        )
        fig.tight_layout(rect=[0, 0.2, 0.82, 0.96])
        sleep_profile_plot = sleep_dir / f"SleepProfile_AllGenotypes_{base}.png"
        fig.savefig(sleep_profile_plot, dpi=300, bbox_inches="tight")

        fig = Figure(figsize=(14, 6))
        FigureCanvasAgg(fig)
        ax = fig.subplots()
        _add_day_night_shading(ax, ld_transitions, use_wc_sleep)
        for idx, (genotype, n_flies, _mean_sleep) in enumerate(combined_series):
            wells = [well for well in mapping.get(genotype, []) if well in binned_sleep_minutes.columns]
            if not wells:
                continue
            geno_sleep_min = binned_sleep_minutes[wells]
            pct_asleep = (geno_sleep_min.sum(axis=1).values / (30.0 * max(n_flies, 1))) * 100.0
            ax.plot(
                sleep_x_vals,
                np.clip(pct_asleep, 0, 100),
                linewidth=2.0,
                color=matplotlib.colormaps["tab10"](idx % 10),
                label=f"{genotype} (n={n_flies})",
            )
        ax.set_ylabel("% Flies Asleep")
        ax.set_ylim(0, 100)
        ax.set_title("% Flies Asleep by Genotype (30-min bins)")
        ax.grid(True, axis="y", alpha=0.3)
        ax.grid(True, axis="x", which="major", alpha=0.4, linewidth=1.2)
        ax.grid(True, axis="x", which="minor", alpha=0.2, linewidth=0.6)
        _format_sleep_xaxis(ax, use_wc_sleep)
        _add_pulse_arrows(ax, sleep_pulses, anchor_sleep, use_wc_sleep, compute_pulse_zt)
        ax.legend(
            title="Genotype",
            loc="center left",
            bbox_to_anchor=(1.01, 0.5),
            fontsize=8,
            title_fontsize=8,
            framealpha=0.9,
            borderpad=0.25,
            labelspacing=0.25,
            handlelength=1.4,
            columnspacing=0.8,
        )
        fig.tight_layout(rect=[0, 0.2, 0.82, 0.96])
        sleep_pct_plot = sleep_dir / f"SleepPctAsleep_AllGenotypes_{base}.png"
        fig.savefig(sleep_pct_plot, dpi=300, bbox_inches="tight")

    return {
        "sleep_totals_workbook": sleep_totals_book,
        "sleep_profile_plot": sleep_profile_plot,
        "sleep_pct_plot": sleep_pct_plot,
        "individual_dir": individual_dir,
    }


def _write_sleep_totals_workbook(
    df: pd.DataFrame,
    mapping: Dict[str, List[str]],
    genotype_order: List[str],
    filename_start_dt: Optional[datetime],
    sleep_per_well: Dict[str, np.ndarray],
    sleep_dir: Path,
    base: str,
) -> Optional[Path]:
    if filename_start_dt is None:
        return None
    runtime_sleep = pd.to_numeric(df["RUNTIME"], errors="coerce")
    rt_valid_mask = runtime_sleep.notna()
    period_df = pd.DataFrame({"RUNTIME": runtime_sleep[rt_valid_mask]}).copy()
    period_df["wall_clock"] = filename_start_dt + pd.to_timedelta(period_df["RUNTIME"], unit="s")
    period_df["Period"] = np.where(
        (period_df["wall_clock"].dt.hour >= 8) & (period_df["wall_clock"].dt.hour < 20),
        "Day",
        "Night",
    )
    period_df["Date"] = period_df["wall_clock"].dt.normalize()
    night_after_midnight = (period_df["Period"] == "Night") & (period_df["wall_clock"].dt.hour < 8)
    period_df.loc[night_after_midnight, "Date"] = period_df.loc[night_after_midnight, "Date"] - pd.to_timedelta(1, unit="D")
    period_df["Date"] = period_df["Date"].dt.strftime("%Y-%m-%d")
    period_df["Period"] = pd.Categorical(period_df["Period"], categories=["Day", "Night"], ordered=True)

    from openpyxl.styles import Border, PatternFill, Side

    workbook_path = sleep_dir / f"SleepTotals_ByGenotype_{base}.xlsx"
    used_sheet_names = set()
    day_fill = PatternFill(fill_type="solid", start_color="FFFDF2", end_color="FFFDF2")
    night_fill = PatternFill(fill_type="solid", start_color="EEF3FF", end_color="EEF3FF")
    period_end_border = Border(bottom=Side(style="thick", color="000000"))

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as totals_writer:
        for genotype in genotype_order:
            wells_for_geno = [well for well in mapping.get(genotype, []) if well in sleep_per_well]
            if not wells_for_geno:
                continue
            well_order = {well: idx for idx, well in enumerate(wells_for_geno)}
            period_keys = (
                period_df[["Date", "Period"]]
                .drop_duplicates()
                .sort_values(["Date", "Period"])
                .reset_index(drop=True)
            )
            key_pairs = list(period_keys.itertuples(index=False, name=None))
            per_well_tables = []
            for well in wells_for_geno:
                sleep_flags = pd.Series(sleep_per_well[well], index=df.index)
                well_df = period_df[["Date", "Period"]].copy()
                well_df["sleep_flag"] = pd.to_numeric(sleep_flags.loc[period_df.index], errors="coerce").fillna(0).astype(int)
                agg = (
                    well_df.groupby(["Date", "Period"], sort=False, observed=False)["sleep_flag"]
                    .sum()
                    .reset_index(name="Total_Sleep_Seconds")
                )
                if key_pairs:
                    agg = pd.DataFrame(key_pairs, columns=["Date", "Period"]).merge(agg, on=["Date", "Period"], how="left")
                    agg["Total_Sleep_Seconds"] = pd.to_numeric(agg["Total_Sleep_Seconds"], errors="coerce").fillna(0).astype(int)
                if agg.empty:
                    continue
                agg["Total_Sleep_Minutes"] = agg["Total_Sleep_Seconds"] / 60.0
                agg.insert(0, "Well_ID", well)
                agg.insert(1, "Genotype", genotype)
                per_well_tables.append(agg[["Well_ID", "Genotype", "Date", "Period", "Total_Sleep_Seconds", "Total_Sleep_Minutes"]])
            if not per_well_tables:
                continue
            genotype_table = pd.concat(per_well_tables, ignore_index=True)
            genotype_table["Well_Order"] = genotype_table["Well_ID"].map(well_order)
            genotype_table["Period"] = pd.Categorical(genotype_table["Period"], categories=["Day", "Night"], ordered=True)
            genotype_table = genotype_table.sort_values(["Date", "Period", "Well_Order"]).drop(columns=["Well_Order"]).reset_index(drop=True)
            sheet_name = sanitize_excel_sheet_name(genotype, used_sheet_names)
            genotype_table.to_excel(totals_writer, sheet_name=sheet_name, index=False)
            worksheet = totals_writer.sheets[sheet_name]
            end_rows = genotype_table.groupby(["Date", "Period"], sort=False, observed=False).size().cumsum().tolist()
            for data_row_end in end_rows:
                excel_row = int(data_row_end) + 1
                for col_idx in range(1, len(genotype_table.columns) + 1):
                    worksheet.cell(row=excel_row, column=col_idx).border = period_end_border
            for data_row_idx, period_value in enumerate(genotype_table["Period"], start=2):
                row_fill = day_fill if str(period_value) == "Day" else night_fill
                for col_idx in range(1, len(genotype_table.columns) + 1):
                    worksheet.cell(row=data_row_idx, column=col_idx).fill = row_fill
    return workbook_path


def _compute_sleep_binary(activity_values, threshold: int = 300) -> np.ndarray:
    inactive = (np.asarray(activity_values) == 0).astype(np.int8)
    padded = np.concatenate([[0], inactive, [0]])
    diffs = np.diff(padded)
    run_starts = np.where(diffs == 1)[0]
    run_ends = np.where(diffs == -1)[0]
    sleep = np.zeros(len(inactive), dtype=np.int8)
    for run_start, run_end in zip(run_starts, run_ends):
        if (run_end - run_start) >= threshold:
            sleep[run_start:run_end] = 1
    return sleep


def _collect_light_dark_segments(df: pd.DataFrame, anchor_sleep: Optional[datetime], use_wc_sleep: bool):
    transitions = []
    if "L_OR_D" not in df.columns:
        return transitions
    local = df[["RUNTIME", "L_OR_D"]].copy()
    local["RUNTIME"] = pd.to_numeric(local["RUNTIME"], errors="coerce")
    local = local.dropna(subset=["RUNTIME"]).sort_values("RUNTIME")
    local["is_light"] = local["L_OR_D"].apply(_to_is_light)
    if use_wc_sleep and anchor_sleep is not None:
        local["x"] = [anchor_sleep + timedelta(seconds=float(runtime)) for runtime in local["RUNTIME"]]
    else:
        local["x"] = local["RUNTIME"].values / 3600.0
    xs = local["x"].tolist()
    states = local["is_light"].tolist()
    if xs:
        seg_start, seg_state = xs[0], states[0]
        for idx in range(1, len(xs)):
            if states[idx] != seg_state:
                if seg_state is not None:
                    transitions.append((seg_start, xs[idx], seg_state))
                seg_start, seg_state = xs[idx], states[idx]
        if seg_state is not None:
            transitions.append((seg_start, xs[-1], seg_state))
    return transitions


def _add_day_night_shading(ax, ld_transitions, use_wc_sleep: bool):
    for seg_start, seg_end, is_light in ld_transitions:
        color = "#ffe066" if is_light else "#001f3f"
        ax.axvspan(seg_start, seg_end, facecolor=color, alpha=0.18, zorder=0)
        try:
            mid = seg_start + (seg_end - seg_start) / 2 if use_wc_sleep else (seg_start + seg_end) / 2
            ax.text(mid, 0.98, "L" if is_light else "D", transform=ax.get_xaxis_transform(), ha="center", va="top", fontsize=12, alpha=0.9)
        except Exception:
            pass


def _add_pulse_arrows(ax, sleep_pulses, anchor_sleep: Optional[datetime], use_wc_sleep: bool, compute_pulse_zt):
    if not sleep_pulses:
        return
    x0_num, x1_num = ax.get_xlim()
    for pulse_idx, (start_row, _end_row, _day, p_hour, _itemp, start_runtime) in enumerate(sleep_pulses, start=1):
        if pd.isna(start_runtime):
            continue
        if use_wc_sleep and anchor_sleep is not None:
            pulse_x = anchor_sleep + timedelta(seconds=float(start_runtime))
            label_time = pulse_x.strftime("%I:%M %p").lstrip("0")
            px_num = mdates.date2num(pulse_x)
        else:
            pulse_x = float(start_runtime) / 3600.0
            label_time = f"{pulse_x:.1f}h"
            px_num = pulse_x
        pulse_zt = compute_pulse_zt(start_runtime, p_hour, start_row)
        label_parts = [f"P#{pulse_idx}", label_time]
        if pulse_zt is not None:
            label_parts.append(f"ZT{int(pulse_zt)}")
        frac = ((px_num - x0_num) / (x1_num - x0_num) if (x1_num - x0_num) else 0.5)
        dx = 12 if frac < 0.08 else (-12 if frac > 0.92 else 0)
        dy = [-50, -70, -90][(pulse_idx - 1) % 3]
        ax.annotate(
            "\n".join(label_parts),
            xy=(pulse_x, -0.02),
            xycoords=("data", "axes fraction"),
            xytext=(dx, dy),
            textcoords="offset points",
            ha="center",
            va="top",
            fontsize=8,
            clip_on=False,
            arrowprops=dict(arrowstyle="-|>", lw=0.8, color="black"),
        )


def _format_sleep_xaxis(ax, use_wc_sleep: bool):
    if use_wc_sleep:
        _apply_wall_clock_xaxis(ax)
        ax.set_xlabel("Wall Clock Time")
    else:
        ax.set_xlabel("Time (hours from start)")


def _apply_wall_clock_xaxis(ax) -> None:
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    ax.xaxis.set_minor_locator(mdates.MinuteLocator(byminute=[0, 30]))

    def _major_fmt(value, _pos):
        dt_val = mdates.num2date(value)
        if getattr(dt_val, "tzinfo", None) is not None:
            dt_val = dt_val.replace(tzinfo=None)
        return dt_val.strftime("%I %p").lstrip("0")

    ax.xaxis.set_major_formatter(FuncFormatter(_major_fmt))
    ax.xaxis.set_minor_formatter(NullFormatter())
    _mpl_setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right", rotation_mode="anchor")
    ax.tick_params(axis="x", which="major", length=8, width=1.2)
    ax.tick_params(axis="x", which="minor", length=4, width=0.6)


def _to_is_light(value):
    if pd.isna(value):
        return None
    try:
        return float(value) != 0
    except Exception:
        text = str(value).strip().lower()
        if text in {"l", "light", "day", "1", "true", "yes"}:
            return True
        if text in {"d", "dark", "0", "false", "no"}:
            return False
        return None
